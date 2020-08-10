#! python3
#  sendDuesReminder.py - Sends reminder to those who haven't paid for last month.
#                       The necessary data are included in Excel file.

import openpyxl, smtplib

wb = openpyxl.load_workbook('duesRecords.xlsx')
ws = wb.active

lastCol = ws.max_column
lastMonth = ws.cell(row=1, column=lastCol).value

# Check each member's payment status:
unpaidMembers = {}
for i in range(2, ws.max_row+1):
    payment = ws.cell(i, lastCol).value
    if payment != 'paid':
        name = ws.cell(i, 1).value
        email = ws.cell(i, 2).value
        unpaidMembers[name] = email

# Login to email account:
smtpObj = smtplib.SMTP_SSL('smtp.poczta.example.pl', 465)
smtpObj.ehlo()
smtpObj.login('example@example.pl', 'somepassword')

# Sending reminders to obtained list of members:
for name, email in unpaidMembers.items():
    body = '''
Subject: %s dues unpaid.\nDear %s,\n Records show that you have not
paid dues for %s. Please make this payment as soon as possible to help us
close the period. Thank you, Bro!''' % (lastMonth, name, lastMonth)

    print('Sending message to %s...' % (name))
    sendmailStatus = smtpObj.sendmail('example@example.pl', email, body)

    if sendmailStatus != {}:
        print('There was a problem with sending email to %s: %s' %
              (email, sendmailStatus))
smtpObj.quit()
