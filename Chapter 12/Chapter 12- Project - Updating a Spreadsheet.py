#! python3
# Updating a Spreadsheet

import openpyxl
from openpyxl.styles import Font, NamedStyle #formatting excel cells
                    #it allows type 'Font()' instead of 'openpyxl.styles.Font()'

wb = openpyxl.load_workbook('produceSales.xlsx', data_only=True)
        # 'data_only=True' the value of formulas will be shown
        # 'data_only=False' the formula of the cell will be shown
# 'load_workbook' doesn't load 'charts'
print(wb.sheetnames)
sheet = wb['Sheet']

PRICE_UPDATES = { 'Garlic': 3.07,
                  'Celery': 1.19,
                  'Lemon': 1.27 }

#Loop through the rows and update the prices:
for rowNum in range(2, sheet.max_row): # the row 1. is a header.
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('produceSales_updated.xlsx')

# Read data from one spreadsheet and write it to parts of other spreadsheets.
# Read data from websites, text files, or the clipboard and write it to a
    # spreadsheet.
# Automatically "clean up" data in spreadsheets. E.x. it could use
    # Regular Expressions to read multiple formats of phone number
    # and edit them to a single, standard format.


#Setting the Font Style of Cells:
myFont = Font(size=24, italic=True, bold=True)
myStyle = NamedStyle(name='myStyle')
myStyle.font = myFont
for cell in sheet['1']:
    cell.style = myStyle    



#Formulas:
sheet['E2'] = '=SUM(D:D)'
print(sheet['E2']) #to get value instead of text of formula I should
    # load_workbook with feature 'data_only=True'

#Setting Raw and Column Width:
## sheet.raw_dimentions[row].height = ..
## sheet.column_dimentions['B'].width = ..

#Mergin and Unmerging Cells:
## 'merge_cells()' Sheet Method
## 'unmerge_cells()' Sheet Method

#Freeze Panes:
## 'freeze_panes' Attribute
## set 'freeze_panes=None' to unfreeze all panes
sheet.freeze_panes = 'A2' # to freeze first raw

#Charts:
'''
    1. Create a 'Reference' object from a rectangular selection of cells
    2. Create a 'Series' object by passing in the 'Reference' object
    3. Create a 'Chart' object
    4. Append the 'Series' object to the 'Chart' object
    5. (optionally) set the 'drawing.top, drawing.left, drawing.width,
        drawing.height' variables of the 'Chart' object
    6. Add the 'Chart' object to the 'Worksheet' object.
                                                                         '''

refObj    = openpyxl.chart.Reference(sheet,
                min_col=4, min_row=2, max_col=4, max_row=500)
seriesObj = openpyxl.chart.Series(refObj, title='Series one')
chartObj  = openpyxl.chart.LineChart()
                        # .BarChart
                        # .ScatterChart
                        # .PieChart
chartObj.append(seriesObj)
sheet.add_chart(chartObj, "F4")
wb.save('produceSales_updatedStyles.xlsx')
