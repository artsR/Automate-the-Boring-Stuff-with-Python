#! python3
# multiplicationTable.py - Takes the number as an argument and creates an NxN
#   multiplication table in an Excel Document

import openpyxl
from openpyxl.styles import Font

n = int(input('Input size of matrix...'))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = str('Multiplication %sx%s' % (n, n))
for i in range(2,n+2):
    ws.cell(row=i, column=1).value = i-1
    ws.cell(row=i, column=1).font = Font(bold=True)
    for j in range(2,n+2):
        ws.cell(row=1, column=j).value = j-1
        ws.cell(row=1, column=j).font = Font(bold=True)
        ws.cell(row=i, column=j).value = (i-1)*(j-1)

wb.save('multiplicationTable.xlsx')

