#! python3
#

import openpyxl
from openpyxl.utils import get_column_letter

filename = 'multiplicationTable.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active

'''
print(ws[get_column_letter(2)])
print(ws['A'])
print(ws[1])
print(ws.columns)
print(ws.rows)
print(list(ws.rows))
print(ws['A2'].value)
print(ws.cell(1,1).value)
print(tuple(ws['A1':'C3']))
'''
wb2 = openpyxl.Workbook()
ws2 = wb2.active

for row in range(1, ws.max_row+1):
   for col in range(1, ws.max_column+1):
        ws2.cell(col,row).value = ws.cell(row,col).value
        

wb2.save('Chapter 12 - Taks3 - Cell Inverter.xlsx')
