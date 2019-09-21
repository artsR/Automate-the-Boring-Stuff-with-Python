######### CHAPTER 12 ##########

import openpyxl

# Creating and Saving Excel Document
## .get_sheet_names()  -> lists names of all sheets in workbook.
## .get_active_sheet() -> points the active sheet
## .save('name.xlsx')  -> saves dokument
wb = openpyxl.Workbook()
print(wb.get_sheet_names())
sheet = wb.get_active_sheet()
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.get_sheet_names())

'''
print(ws[get_column_letter(2)])
print(ws['A'])
print(ws[1])
print(ws.columns)
print(ws.rows)
print(list(ws.rows))
print(ws['A2'].value)
print(ws.cell(1,1).value)
print(tuple(ws['A1':'C3']))
'''

#wb2 = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
sheet.title = 'Spam spam spam'
wb.save('example_sheet_copy.xlsx') # The change in the workbook won't be saved until
            # I use 'save' Method.

# Creating and Removing Sheets
## .get_sheet_by_name('name') -> points specific sheet in workbook.
## .create_sheet() Method
## .remove_sheet() Method
print(wb.get_sheet_names())
wb.create_sheet(index=0, title='First Sheet') # creates a new Worksheet object
print(wb.get_sheet_names())
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
wb.remove_sheet(wb.get_sheet_by_name('Spam spam spam'))
print(wb.get_sheet_names())

# Writing Values to Cells
wb3 = openpyxl.Workbook()
print(wb3.get_sheet_names())
sheet = wb3.get_sheet_by_name('Sheet')
sheet['A1'] = 'Hello, Sir!'
print(sheet['A1'].value)
wb3.save('example_sheet.xlsx')

