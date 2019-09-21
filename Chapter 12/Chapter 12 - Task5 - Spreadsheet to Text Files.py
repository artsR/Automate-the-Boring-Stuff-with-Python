#! python3
# Spreadsheet to Text Files
# Writes the cells of each column into separate text file.

import openpyxl, logging
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('Chapter 12 - TEXTtoEXCEL.xlsx')
ws = wb.active

for col in range(1, ws.max_column+1):
    content = []
    textFile = open('EXCELtoTEXT_%s.txt' % col, 'w')
    for row in range(1, len(ws[get_column_letter(col)])+1):
        content.append(ws.cell(row, col).value)
        '''try:
            textFile.write(ws.cell(row, col).value)
        except:
            pass'''
    textFile.write(''.join(filter(None, content)))
    textFile.close()

print('Saving to files completed.')
    
