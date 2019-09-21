########### CHAPTER 12 ############

import openpyxl # it's Module which allows work with Excel and LibreOffice as well
                # http://openpyxl.readthedocs.org/

'''
wb = openpyxl.load_workbook('example.xlsx') # load Excel document into the variable
type(wb)
        # 'data_only=True' the value of formulas will be shown
        # 'data_only=False' the formula of the cell will be shown
        
wb.get_sheet_names() # list of all sheet in the workbook
sheet1 = wb.get_sheet_by_name('Sheet2') #wb['Sheet2']
sheet1
type(sheet1)   # Worksheet Object
sheet.title    # sheet's name
active_sheet = wb.get_active_sheet()

sheet['A1'].value   # value of cell 'A1'
wydatki = sheet['B1'] #(1) alternatively I can also use cell() method:
                        # sheet.cell(row=1, column=2)
wydatki.value
'Row ' + str(wydatki.row) + ', Column ' + wydatki.column + ' is ' + wydatki.value
# 'coordinate' returns number of cell e.x. "B1"
'Cell ' + wydatki.coordinate + ' is ' + wydatki.value
'''

# OpenPyXL automacally interpret the dates in column and return them properly
# to the value type e.x. date will be returned as datetime values and not just
# string.

#sheet.cell(row=1, column=2).value # (1)
'''
for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)
                                                '''
# Obtaining size of the worksheet:
'''
sheet1.get_highest_row()
sheet1.get_highest_column() # returns 'integer' number, not letter.
'''

# Convert from numbers to letters
from openpyxl.cell import get_column_letter, column_index_from_string
#openpyxl.cell.get_column_letter(2) # returns 'B'
get_column_letter(2)
get_column_letter(sheet1.get_highest_column())
# Convert from letters to numbers
column_index_from_string('B') # returns '2'

# Getting all 'Cell Objects' in specific area:
tuple(sheet1['A1':'C3']) #it contains three tuples: one for each row.
for rowObj in sheet1['A1':'C3']:
    for cellObj in rowObj:
        print(cellObj.coordinate, cellObj.value)
    print('--- END of ROW ---')

# Another way to go through cells:
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()
sheet.column[1] # returns the column 'B'. ([0] returns column 'A')
for cellObj in sheet.column[1]:
    print(cellObj.value)







